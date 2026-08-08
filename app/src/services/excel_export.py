# std
from datetime import datetime, timedelta
from io import BytesIO

# openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# sqlalchemy
from sqlalchemy import func
from sqlalchemy.orm import Session

# app
from app.core.constants import mexico_now
from app.src.models import Order, OrderDetail, Product, Sale, SaleDetail
from app.src.providers.report import ReportProvider

# --- estilos ---
_BRAND = "16A34A"          # verde
_HEADER_BG = PatternFill("solid", fgColor="1F2937")   # gris oscuro
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=16, color="111827")
_SUB_FONT = Font(size=10, color="6B7280")
_TOTAL_FONT = Font(bold=True, size=11)
_TOTAL_BG = PatternFill("solid", fgColor="E5E7EB")
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_WRAP = Alignment(wrap_text=True, vertical="top")  # una línea por producto en la celda
_TOP = Alignment(vertical="top")
_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_MONEY = '"$"#,##0.00'
_QTY = "#,##0.###"


def _style_header(ws: Worksheet, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_BG
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sheet_title(ws: Worksheet, title: str, subtitle: str, span: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    ws.cell(row=2, column=1, value=subtitle).font = _SUB_FONT
    ws.row_dimensions[1].height = 24


def _period_bounds():
    now = mexico_now()
    today = now.date()
    day_start = datetime(today.year, today.month, today.day)
    day_end = day_start + timedelta(days=1)
    week_start = datetime.combine(today - timedelta(days=today.weekday()), datetime.min.time())
    month_start = datetime(today.year, today.month, 1)
    return day_start, day_end, week_start, month_start


def _section_band(ws: Worksheet, row: int, label: str, span: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = _LEFT
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=_BRAND)


def _customer_names(db: Session) -> dict[int, str]:
    from app.src.models import Customer

    return {c.id: c.customer_name for c in db.query(Customer).all()}


def _sales_sheet(ws: Worksheet, db: Session, names: dict[int, str]) -> None:
    _sheet_title(ws, "Ventas de mostrador", "Por temporalidad: hoy, semana y mes actual", 5)
    headers = ["#", "Fecha", "Cliente", "Productos", "Total"]
    day_start, day_end, week_start, month_start = _period_bounds()
    sections = [
        ("HOY", (Sale.date >= day_start) & (Sale.date < day_end)),
        ("SEMANA ACTUAL", Sale.date >= week_start),
        ("MES ACTUAL", Sale.date >= month_start),
    ]

    r = 3
    for label, sale_filter in sections:
        r += 1
        _section_band(ws, r, label, len(headers))
        r += 1
        for i, h in enumerate(headers, start=1):
            ws.cell(row=r, column=i, value=h)
        _style_header(ws, r, len(headers))

        sales = db.query(Sale).filter(sale_filter).order_by(Sale.date.desc()).all()
        subtotal = 0.0
        if not sales:
            r += 1
            ws.cell(row=r, column=1, value="Sin ventas en este periodo").font = _SUB_FONT
        for sale in sales:
            r += 1
            # Un producto por línea dentro de la misma celda (se ve como "filitas")
            productos = "\n".join(
                f"{d.product.name} ×{d.quantity:g}" for d in sale.sales_details
            )
            ws.cell(row=r, column=1, value=sale.id).alignment = _CENTER
            date_cell = ws.cell(
                row=r, column=2, value=sale.date.strftime("%d/%m/%Y %H:%M") if sale.date else ""
            )
            date_cell.alignment = _TOP
            ws.cell(row=r, column=3, value=names.get(sale.customer_id, "—")).alignment = _TOP
            prod_cell = ws.cell(row=r, column=4, value=productos)
            prod_cell.alignment = _WRAP
            total_cell = ws.cell(row=r, column=5, value=sale.total or 0.0)
            total_cell.number_format = _MONEY
            total_cell.alignment = _TOP
            # Altura acorde al número de productos (para que se vean todas las líneas)
            lines = max(1, len(sale.sales_details))
            ws.row_dimensions[r].height = 15 * lines
            subtotal += sale.total or 0.0

        r += 1
        label_cell = ws.cell(row=r, column=4, value="Subtotal")
        label_cell.font = _TOTAL_FONT
        label_cell.alignment = _RIGHT
        sub = ws.cell(row=r, column=5, value=subtotal)
        sub.number_format = _MONEY
        sub.font = _TOTAL_FONT
        for c in (4, 5):
            ws.cell(row=r, column=c).fill = _TOTAL_BG
        r += 1  # separación entre secciones

    _autosize(ws, [6, 18, 26, 46, 14])


def _orders_sheet(ws: Worksheet, db: Session, names: dict[int, str]) -> None:
    _sheet_title(ws, "Pedidos", "Por temporalidad: hoy, semana y mes actual", 8)
    headers = ["#", "Fecha", "Cliente", "Total", "Pagado", "Restante", "Entrega", "Pago"]
    day_start, day_end, week_start, month_start = _period_bounds()
    sections = [
        ("HOY", (Order.date >= day_start) & (Order.date < day_end)),
        ("SEMANA ACTUAL", Order.date >= week_start),
        ("MES ACTUAL", Order.date >= month_start),
    ]

    r = 3
    for label, order_filter in sections:
        r += 1
        _section_band(ws, r, label, len(headers))
        r += 1
        for i, h in enumerate(headers, start=1):
            ws.cell(row=r, column=i, value=h)
        _style_header(ws, r, len(headers))

        orders = db.query(Order).filter(order_filter).order_by(Order.date.desc()).all()
        s_total = s_paid = 0.0
        if not orders:
            r += 1
            ws.cell(row=r, column=1, value="Sin pedidos en este periodo").font = _SUB_FONT
        for o in orders:
            r += 1
            paid = o.amount_paid or 0.0
            remaining = max((o.total or 0.0) - paid, 0.0)
            ws.cell(row=r, column=1, value=o.id).alignment = _CENTER
            ws.cell(row=r, column=2, value=o.date.strftime("%d/%m/%Y %H:%M") if o.date else "")
            ws.cell(row=r, column=3, value=names.get(o.customer_id, "—"))
            for col, val in ((4, o.total or 0.0), (5, paid), (6, remaining)):
                cell = ws.cell(row=r, column=col, value=val)
                cell.number_format = _MONEY
            ws.cell(row=r, column=7, value=o.status.capitalize() if o.status else "")
            ws.cell(row=r, column=8, value=o.payment_status)
            s_total += o.total or 0.0
            s_paid += paid

        r += 1
        label_cell = ws.cell(row=r, column=3, value="Subtotal")
        label_cell.font = _TOTAL_FONT
        label_cell.alignment = _RIGHT
        for col, val in ((4, s_total), (5, s_paid), (6, max(s_total - s_paid, 0.0))):
            cell = ws.cell(row=r, column=col, value=val)
            cell.number_format = _MONEY
            cell.font = _TOTAL_FONT
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = _TOTAL_BG
        r += 1  # separación entre secciones

    _autosize(ws, [6, 18, 26, 14, 14, 14, 14, 16])


def _agg_products(db: Session, sale_filter=None, order_filter=None) -> dict[str, list[float]]:
    """Cantidad e ingreso por producto (ventas + pedidos), con filtro opcional de fecha."""
    agg: dict[str, list[float]] = {}  # name -> [qty, income]

    def add(rows):
        for name, qty, income in rows:
            entry = agg.setdefault(name, [0.0, 0.0])
            entry[0] += qty or 0.0
            entry[1] += income or 0.0

    sq = (
        db.query(Product.name, func.sum(SaleDetail.quantity), func.sum(SaleDetail.subtotal))
        .join(SaleDetail, SaleDetail.product_id == Product.id)
        .join(Sale, Sale.id == SaleDetail.sale_id)
    )
    if sale_filter is not None:
        sq = sq.filter(sale_filter)
    add(sq.group_by(Product.id).all())

    oq = (
        db.query(Product.name, func.sum(OrderDetail.quantity), func.sum(OrderDetail.subtotal))
        .join(OrderDetail, OrderDetail.product_id == Product.id)
        .join(Order, Order.id == OrderDetail.order_id)
    )
    if order_filter is not None:
        oq = oq.filter(order_filter)
    add(oq.group_by(Product.id).all())

    return agg


def _products_sheet(ws: Worksheet, db: Session) -> None:
    _sheet_title(
        ws,
        "Productos vendidos",
        "Cantidad e ingreso por producto (ventas + pedidos): hoy, semana y mes actual",
        7,
    )

    day_start, day_end, week_start, month_start = _period_bounds()

    periods = [
        ("Hoy", (Sale.date >= day_start) & (Sale.date < day_end),
                (Order.date >= day_start) & (Order.date < day_end)),
        ("Semana actual", Sale.date >= week_start, Order.date >= week_start),
        ("Mes actual", Sale.date >= month_start, Order.date >= month_start),
    ]
    data = [(label, _agg_products(db, sf, of)) for label, sf, of in periods]

    # Encabezado agrupado: fila 4 (grupo) + fila 5 (Cant / Ingreso)
    g_row, h_row = 4, 5
    ws.cell(row=g_row, column=1, value="Producto")
    ws.merge_cells(start_row=g_row, start_column=1, end_row=h_row, end_column=1)
    col = 2
    for label, _ in data:
        ws.merge_cells(start_row=g_row, start_column=col, end_row=g_row, end_column=col + 1)
        ws.cell(row=g_row, column=col, value=label)
        ws.cell(row=h_row, column=col, value="Cant.")
        ws.cell(row=h_row, column=col + 1, value="Ingreso")
        col += 2
    _style_header(ws, g_row, 7)
    _style_header(ws, h_row, 7)

    # Orden por ingreso del mes actual, de mayor a menor
    month_agg = data[-1][1]
    names = sorted(month_agg.keys(), key=lambda n: month_agg[n][1], reverse=True)

    r = h_row
    for name in names:
        r += 1
        ws.cell(row=r, column=1, value=name)
        col = 2
        for _, agg in data:
            qty, income = agg.get(name, (0.0, 0.0))
            ws.cell(row=r, column=col, value=qty).number_format = _QTY
            ws.cell(row=r, column=col + 1, value=income).number_format = _MONEY
            col += 2

    # Fila de totales
    r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = _TOTAL_FONT
    col = 2
    for _, agg in data:
        tot_income = sum(v[1] for v in agg.values())
        ws.cell(row=r, column=col).fill = _TOTAL_BG
        cell = ws.cell(row=r, column=col + 1, value=tot_income)
        cell.number_format = _MONEY
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_BG
        col += 2
    ws.cell(row=r, column=1).fill = _TOTAL_BG

    _autosize(ws, [28, 9, 13, 9, 13, 9, 13])
    ws.freeze_panes = "B6"


def _summary_sheet(ws: Worksheet, db: Session) -> None:
    rp = ReportProvider(db)
    today = rp.today_summary()
    monthly = rp.monthly_income()
    losses = rp.losses_total()
    fin = rp.finance()

    _sheet_title(
        ws,
        "Resumen del negocio",
        f"Generado el {mexico_now().strftime('%d/%m/%Y %H:%M')}",
        2,
    )

    def section(row: int, title: str) -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_BRAND)
        cell.alignment = _LEFT
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=_BRAND)
        return row + 1

    def kv(row: int, label: str, value, money: bool = False) -> int:
        ws.cell(row=row, column=1, value=label).font = Font(color="374151")
        cell = ws.cell(row=row, column=2, value=value)
        cell.alignment = _RIGHT
        cell.font = Font(bold=True)
        if money:
            cell.number_format = _MONEY
        return row + 1

    r = 4
    r = section(r, "Hoy")
    r = kv(r, "Ingresos de hoy", today["income_total"], money=True)
    r = kv(r, "Ventas de mostrador", today["sales_count"])
    r = kv(r, "Pedidos completados", today["orders_completed_count"])
    r = kv(r, "Kilos de tortilla devueltos", losses["today"])
    r += 1

    r = section(r, "Mes")
    r = kv(r, "Ingresos del mes", monthly["income_total"], money=True)
    r = kv(r, "Pedidos del mes", monthly["orders_count"])
    r = kv(r, "Kilos de tortilla devueltos", losses["month"])
    r += 1

    r = section(r, "Desde tu última compra de insumos")
    if fin["income_since"]:
        r = kv(r, "Fecha de la última compra", fin["income_since"])
        r = kv(r, "Gasto en insumos", fin["total_expense"], money=True)
        r = kv(r, "Has ganado desde entonces", fin["income"], money=True)
        r = kv(r, "Diferencia", fin["net"], money=True)
    else:
        r = kv(r, "Sin compras de insumos recientes", "")

    _autosize(ws, [32, 20])


def build_sales_workbook(db: Session) -> bytes:
    """Genera un Excel con formato: Resumen, Ventas, Pedidos y Productos."""
    wb = Workbook()
    names = _customer_names(db)

    ws_summary = wb.active
    ws_summary.title = "Resumen"
    _summary_sheet(ws_summary, db)

    _sales_sheet(wb.create_sheet("Ventas"), db, names)
    _orders_sheet(wb.create_sheet("Pedidos"), db, names)
    _products_sheet(wb.create_sheet("Productos"), db)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
